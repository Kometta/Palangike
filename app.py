from flask import Flask, render_template, redirect, url_for, request
import json
import datetime
import time

import os
import os.path
import pandas as pd
import csv
import mysql.connector

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

import openpyxl
from pathlib import Path

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/classroom.courses.readonly',
          'https://www.googleapis.com/auth/classroom.student-submissions.students.readonly',
          "https://www.googleapis.com/auth/classroom.profile.emails",
          "https://www.googleapis.com/auth/classroom.rosters"]

app = Flask(__name__)

THIS_FOLDER = Path(__file__).parent.resolve()

app.config['UPLOAD_FOLDER'] = os.path.join(THIS_FOLDER, 'uploads')
app.config['DOWNLOAD_FOLDER'] = os.path.join(THIS_FOLDER, 'downloads')


def load_config():
    with open(os.path.join(THIS_FOLDER, 'config.json')) as config_file:
        result = json.load(config_file)
    return result


config = load_config()

DB_USERNAME = config.get('DB_USER')
DB_PASSWORD = config.get('DB_PASS')
DB_DATABASE = config.get('DB_NAME')
DB_DATAHOST = config.get('DB_HOST')


def extract_data_from_excel(filename):
    wb = openpyxl.load_workbook(filename)
    data_dict = {}

    for sheet_name in wb.sheetnames:
        if sheet_name != 'Final Scores':
            continue
        sheet = wb[sheet_name]
        sheet_data = []

        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(row)

        data_dict[sheet_name] = sheet_data

    return data_dict


def get_classroom_data(service):
    courseList = classroom_get_course_ids(service)

    # print_list(courseList)

    chosenCourseId = "545328896162"

    courseWorkList = classroom_get_courseWork_ids(service, chosenCourseId)

    # print_list(courseWorkList)

    chosenCourseWorkId = "601002302667"

    students = get_course_students(service, chosenCourseId)
    grades = get_courseWork_grades(service, chosenCourseId, chosenCourseWorkId)

    save_students_grades(students, grades, chosenCourseWorkId)


def classroom_get_course_ids(service):
        
    # Call the Classroom API
    results = service.courses().list(teacherId='me').execute()
    courses = results.get('courses')

    if not courses:
        print('No courses found.')
        return

    courseList = list()

    for course in courses:
        course_id = course['id']
        course_name = course['name']
        courseList.append({course_id: course_name})

    return courseList


def print_list(idList):
    for item in idList:
        print(item)


def classroom_get_courseWork_ids(service, chosenCourseId):
        

    results = service.courses().courseWork().list(courseId=chosenCourseId, pageSize=10).execute()
    coursesWorks = results["courseWork"]

    if not coursesWorks:
        print('No courses found.')
        return

    courseWorkList = list()

    for work in coursesWorks:
        # Extract the relevant course information
        work_id = work['id']
        work_name = work['title']
        courseWorkList.append({work_id: work_name})

    return courseWorkList


def get_course_students(service, chosenCourseId):
        
    users = service.courses().students().list(courseId=chosenCourseId).execute()["students"]

    students = dict()

    for user in users:
        id = user["userId"]
        name = user["profile"]["name"]["fullName"]
        email = user["profile"]["emailAddress"]
        username = email[:6]

        students[id] = {"name": name, "email": email, "username": username}

    return students


def get_courseWork_grades(service, chosenCourseId, chosenCourseWorkId):
    grades = dict()

    # Call the API to get the list of coursework for the current course
    result = service.courses().courseWork().\
        studentSubmissions().list(courseId=chosenCourseId, courseWorkId=chosenCourseWorkId).execute()

    submissions = result["studentSubmissions"]

    for submission in submissions:
        userId = submission["userId"]
        if 'assignedGrade' not in submission:
            finalScore = "No_score"
            grades[userId] = finalScore
            continue
        finalScore = submission["assignedGrade"]
        grades[userId] = finalScore

    return grades


def save_students_grades(students, grades, chosenCourseWorkId):

    for id in students:
        students[id]["grade"] = grades[id]

    output_file = f'courseWork_{chosenCourseWorkId}.json'

    with open(os.path.join(app.config['DOWNLOAD_FOLDER'], output_file), 'w') as file:
        json.dump(students, file, indent=4)


def process_data(data):
    
    data = list(data.values())[0]
    del data[:2]

    df = pd.DataFrame(data)
    df = df.iloc[:, [1, 2]]
    df = df.drop(0)
    df.columns = ["username", "kahoot_score"]

    df["username"] = transform_username(df['username'])

    df.to_json(os.path.join(app.config['UPLOAD_FOLDER'], 'kahoot_info.json'), orient='records')

    return 


def transform_username(column):
    transformed_column = column.apply(lambda x: x[:6].lower())
    return transformed_column


def merge_data():

    # Get the list of courseWork files in the current directory
    classroom_files = [file for file in os.listdir(app.config['DOWNLOAD_FOLDER']) if file.startswith("courseWork")]

    # Sort the courseWork files by modified time (most recent first)
    sorted_files = sorted(classroom_files, key=lambda x: os.path.getmtime(os.path.join(app.config['DOWNLOAD_FOLDER'], x)), reverse=True)

    # Check if any courseWork files exist
    if sorted_files:
        # Read the contents of the latest JSON file
        courseWorkFilename = sorted_files[0]

    # courseWorkFilename = 'courseWork_'

    classroomData = pd.read_json(os.path.join(app.config['DOWNLOAD_FOLDER'], courseWorkFilename))

    classroomData = classroomData.transpose()
    kahootData = pd.read_json(os.path.join(app.config['DOWNLOAD_FOLDER'], "kahoot_info.json"))
    merged_df = classroomData.merge(kahootData, on='username')

    merged_df.to_json(os.path.join(app.config['DOWNLOAD_FOLDER'], "merged.json"))
    merged_df.to_csv(os.path.join(app.config['DOWNLOAD_FOLDER'], "merged.csv"))

    # save stuff to db

    try:

        connection = mysql.connector.connect(
            host=DB_DATAHOST,
            user=DB_USERNAME,
            password=DB_PASSWORD,
            database=DB_DATABASE,
            connect_timeout=10
        )

        if connection.is_connected():
            print("Connected to the database!")

            cursor = connection.cursor()

            # Delete all entries from the merge_data table
            delete_query = "DELETE FROM merge_data"
            cursor.execute(delete_query)
            connection.commit()

            with open(os.path.join(app.config['DOWNLOAD_FOLDER'], "merged.csv"), 'r') as file:
                csv_data = csv.reader(file)
                next(csv_data)  # Skip the header row

                for row in csv_data:
                    query = "INSERT INTO merge_data (id, name, email, username, grade, kahoot_score) VALUES (%s, %s, %s, %s, %s, %s)"
                    values = (int(row[0]), row[1], row[2], row[3], int(row[4]), int(row[5]))
                    cursor.execute(query, values)

            connection.commit()

    except mysql.connector.Error as error:
        print(f"Error connecting to the database: {error}")

    finally:
        if 'connection' in locals() and connection.is_connected():
            connection.close()
            print("Database connection closed.")


def anonymise_showcase_data(data, show_csv):
    for row in data:
        if show_csv:
            row["username"] = row["username"][0]+"*****"
        else:
            row[2] = row[2][0] + "*****"
    return data


@app.route('/')
def index():
    # Get the list of JSON files in the current directory
    json_files = [file for file in os.listdir(app.config['DOWNLOAD_FOLDER']) if file.endswith('.json')]

    # Sort the JSON files by modified time (most recent first)
    sorted_files = sorted(json_files, key=lambda x: os.path.getmtime(os.path.join(app.config['DOWNLOAD_FOLDER'], x)), reverse=True)

    # Check if any JSON files exist
    if sorted_files:
        # Read the contents of the latest JSON file
        latest_file = os.path.join(app.config['DOWNLOAD_FOLDER'], sorted_files[0])
        # with open(latest_file) as file:
        #     data = json.load(file)

        # Get the first element and file date
        first_element = latest_file
        # file_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        file_date = "%s" % time.ctime(os.path.getmtime(latest_file))

        # Render the index template with data
        return render_template('index.html', first_element=first_element, file_date=file_date)
    else:
        # Render the index template with empty data
        return render_template('index.html', first_element=None, file_date=None)


@app.route('/fetch', methods=['POST'])
def fetch():
    # Implement your code to fetch the newest information from the Google Classroom API
    # Update the JSON file with the new information

    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(os.path.join(THIS_FOLDER, 'token.json')):
        creds = Credentials.from_authorized_user_file(os.path.join(THIS_FOLDER, 'token.json'), SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(os.path.join(THIS_FOLDER, 'credentials.json'), SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(os.path.join(THIS_FOLDER, 'token.json'), 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('classroom', 'v1', credentials=creds)

        get_classroom_data(service)

    except HttpError as error:
        print('An error occurred: %s' % error)

    return redirect(url_for('index'))


@app.route('/kahoot')
def kahoot():
    # Get the list of JSON files in the current directory
    json_files = [file for file in os.listdir(app.config['UPLOAD_FOLDER']) if file.endswith('.json')]

    # Sort the JSON files by modified time (most recent first)
    sorted_files = sorted(json_files, key=lambda x: os.path.getmtime(os.path.join(app.config['UPLOAD_FOLDER'], x)), reverse=True)

    # Check if any JSON files exist
    if sorted_files:
        # Read the contents of the latest JSON file
        latest_file = os.path.join(app.config['UPLOAD_FOLDER'], sorted_files[0])

        # Get the first element and file date
        first_element = latest_file
        # file_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        file_date = "%s" % time.ctime(os.path.getmtime(latest_file))

        # Render the index template with data
        return render_template('kahoot.html', first_element=first_element, file_date=file_date)
    else:
        # Render the index template with empty data
        return render_template('kahoot.html', first_element=None, file_date=None)


@app.route('/extract', methods=['POST'])
def extract():

    # Get the list of Excel files in the current directory
    json_files = [file for file in os.listdir(app.config['UPLOAD_FOLDER']) if file.endswith('.xlsx')]

    # Sort the Excel files by modified time (most recent first)
    sorted_files = sorted(json_files, key=lambda x: os.path.getmtime(os.path.join(app.config['UPLOAD_FOLDER'], x)), reverse=True)

    # Check if any Excel files exist
    if sorted_files:
        # Read the contents of the latest JSON file
        filename = sorted_files[0]

        # filename = 'Kahoot video.xlsx'
        data = extract_data_from_excel(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        process_data(data)

        # with open(os.path.join(app.config['DOWNLOAD_FOLDER'], "kahoot_info.json"), 'w') as file:
        #     json.dump(file_info, file, indent=4)

    return redirect(url_for('kahoot'))


@app.route('/merge')
def merge():

    # Get the list of JSON files in the current directory
    json_files = [file for file in os.listdir(app.config['DOWNLOAD_FOLDER']) if file.endswith('.csv')]

    # Sort the JSON files by modified time (most recent first)
    sorted_files = sorted(json_files, key=lambda x: os.path.getmtime(os.path.join(app.config['DOWNLOAD_FOLDER'], x)), reverse=True)

    # Check if any JSON files exist
    if sorted_files:
        # Read the contents of the latest JSON file
        latest_file = os.path.join(app.config['DOWNLOAD_FOLDER'], sorted_files[0])

        # Get the first element and file date
        first_element = latest_file
        # file_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        file_date = "%s" % time.ctime(os.path.getmtime(latest_file))

        # load data from DB

        # Create a list to store the data
        data = []
        show_csv = False

        try:

            # Set up a connection to the MySQL database
            connection = mysql.connector.connect(
                host=DB_DATAHOST,
                user=DB_USERNAME,
                password=DB_PASSWORD,
                database=DB_DATABASE,
                connect_timeout=10
            )

            if connection.is_connected():
                print("Connected to the database!")

                # Create a cursor to execute SQL queries
                cursor = connection.cursor()

                # Execute the SELECT query to fetch data from the table
                query = "SELECT * FROM merge_data"
                cursor.execute(query)

                # Fetch all rows from the result
                rows = cursor.fetchall()

                # Get the column names from the cursor description
                # columns = [column[0] for column in cursor.description]

                # Add the column names as the first row in the data list
                # data.append(columns)

                # Add the rows of data to the list
                # for row in rows:
                #     data.append(row)

                # Convert rows to lists
                data = [list(row) for row in rows]

                # Close the cursor and the database connection
                cursor.close()
                connection.close()

            else:
                with open(latest_file, 'r') as csv_file:
                    reader = csv.DictReader(csv_file)
                    data = list(reader)

                show_csv = True

        except mysql.connector.Error as error:
            print(f"Error connecting to the database: {error}")

            with open(latest_file, 'r') as csv_file:
                reader = csv.DictReader(csv_file)
                data = list(reader)

            show_csv = True

        finally:
            if 'connection' in locals() and connection.is_connected():
                connection.close()
                print("Database connection closed.")
        
        data = anonymise_showcase_data(data, show_csv)

        print(data)
        
        # Render the index template with data
        return render_template('merge.html', first_element=first_element, file_date=file_date, data=data, show_csv=show_csv)
    else:
        # Render the index template with empty data
        return render_template('merge.html', first_element=None, file_date=None)


@app.route('/merged', methods=['POST'])
def merged():

    merge_data()

    return redirect(url_for('merge'))


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        # Check if a file is selected
        if 'file' not in request.files:
            return 'No file selected'

        file = request.files['file']

        # Check if a file is provided
        if file.filename == '':
            return 'No file selected'

        # Save the file to the uploads directory
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))

        return 'File uploaded successfully'

    return render_template('upload.html')


if __name__ == '__main__':
    app.run(debug=True)

